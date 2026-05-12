import argparse
import itertools
import json
import os
import re
from collections import Counter
from datetime import datetime

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment


URL = "https://www.vietlott.vn/vi/trung-thuong/ket-qua-trung-thuong/winning-number-535#top"
HOME = "https://www.vietlott.vn/vi/"
AJAX_URL = (
    "https://www.vietlott.vn/ajaxpro/Vietlott.PlugIn.WebParts.Game535CompareWebPart,"
    "Vietlott.PlugIn.WebParts.ashx"
)
# Token in trang winning-number-535 (JS): ServerSideDrawResult(RenderInfo, '........', ...)
DEFAULT_AJAX_KEY = "64bdd318"
IMPERSONATE = "chrome124"
# Du lieu cong dong (cap nhat hang ngay), dung khi IP GitHub bi Vietlott chan hoan toan.
DEFAULT_JSONL_URLS = (
    "https://raw.githubusercontent.com/vietvudanh/vietlott-data/master/data/power535.jsonl",
    "https://cdn.jsdelivr.net/gh/vietvudanh/vietlott-data@master/data/power535.jsonl",
)


def _headers(referer: str | None) -> dict[str, str]:
    h: dict[str, str] = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        ),
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,image/apng,*/*;q=0.8"
        ),
        "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cache-Control": "max-age=0",
        "Sec-Ch-Ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin" if referer else "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
    }
    if referer:
        h["Referer"] = referer
    return h


def _ajax_key_from_page_html(html: str) -> str | None:
    m = re.search(
        r"Game535CompareWebPart\.ServerSideDrawResult\([^,]+,\s*'([0-9a-f]{8})'",
        html,
        re.I,
    )
    return m.group(1) if m else None


def _ajax_request_body(ajax_key: str, page_index: int = 0) -> dict:
    return {
        "ORenderInfo": {
            "ExtraParam1": "",
            "ExtraParam2": "",
            "ExtraParam3": "",
            "FullPageAlias": "",
            "IsPageDesign": False,
            "OrgPageAlias": "",
            "PageAlias": "",
            "RefKey": "",
            "SiteAlias": "main.vi",
            "SiteId": "main.frontend.vi",
            "SiteLang": "vi",
            "SiteName": "Vietlott",
            "SiteURL": "",
            "System": 1,
            "UserSessionId": "",
            "WebPage": "",
        },
        "Key": ajax_key,
        "GameDrawId": "",
        "ArrayNumbers": [["" for _ in range(35)] for _ in range(5)],
        "CheckMulti": False,
        "PageIndex": page_index,
    }


def _ajaxpro_headers() -> dict[str, str]:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
        "Content-Type": "text/plain; charset=utf-8",
        "X-AjaxPro-Method": "ServerSideDrawResult",
        "Origin": "https://www.vietlott.vn",
        "Referer": "https://www.vietlott.vn/vi/trung-thuong/ket-qua-trung-thuong/winning-number-535",
    }


def _fetch_html_via_ajaxpro(page_url: str, ajax_key: str) -> str:
    """Lay bang ket qua qua AjaxPro (thuong it bi chan hon GET trang day du tren IP datacenter)."""
    page_url = page_url.split("#", 1)[0]
    try:
        from curl_cffi import requests as curl_requests

        session = curl_requests.Session()
        session.get(HOME, impersonate=IMPERSONATE, timeout=30, headers=_headers(None))
        win_html = None
        try:
            win = session.get(
                page_url,
                impersonate=IMPERSONATE,
                timeout=30,
                headers=_headers(HOME),
            )
            if win.ok:
                win_html = win.text
        except Exception:
            win_html = None
        key = (
            _ajax_key_from_page_html(win_html)
            if win_html
            else None
        ) or ajax_key
        body = _ajax_request_body(key)
        resp = session.post(
            AJAX_URL,
            data=json.dumps(body, ensure_ascii=False),
            headers=_ajaxpro_headers(),
            impersonate=IMPERSONATE,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        val = data.get("value") or {}
        if val.get("Error"):
            msg = val.get("InfoMessage") or str(val)
            raise RuntimeError(f"AjaxPro loi: {msg}")
        html = val.get("HtmlContent") or ""
        if not html.strip():
            raise RuntimeError("AjaxPro tra ve HtmlContent rong.")
        return html
    except ImportError:
        session = requests.Session()
        session.get(HOME, timeout=30, headers=_headers(None))
        win_html = None
        try:
            win = session.get(page_url, timeout=30, headers=_headers(HOME))
            if win.ok:
                win_html = win.text
        except Exception:
            win_html = None
        key = (
            _ajax_key_from_page_html(win_html)
            if win_html
            else None
        ) or ajax_key
        body = _ajax_request_body(key)
        resp = session.post(
            AJAX_URL,
            data=json.dumps(body, ensure_ascii=False),
            headers=_ajaxpro_headers(),
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        val = data.get("value") or {}
        if val.get("Error"):
            msg = val.get("InfoMessage") or str(val)
            raise RuntimeError(f"AjaxPro loi: {msg}")
        html = val.get("HtmlContent") or ""
        if not html.strip():
            raise RuntimeError("AjaxPro tra ve HtmlContent rong.")
        return html


def _fetch_page_html(page_url: str) -> str:
    page_url = page_url.split("#", 1)[0]
    try:
        from curl_cffi import requests as curl_requests

        session = curl_requests.Session()
        session.get(HOME, impersonate=IMPERSONATE, timeout=30, headers=_headers(None))
        resp = session.get(
            page_url,
            impersonate=IMPERSONATE,
            timeout=30,
            headers=_headers(HOME),
        )
        resp.raise_for_status()
        return resp.text
    except ImportError:
        session = requests.Session()
        session.get(HOME, timeout=30, headers=_headers(None))
        resp = session.get(page_url, timeout=30, headers=_headers(HOME))
        resp.raise_for_status()
        return resp.text


def parse_args():
    parser = argparse.ArgumentParser(description="Auto update 535.xlsm from Vietlott latest draw.")
    parser.add_argument("--file", default="535.xlsm", help="Path to XLSM file (relative to repo root).")
    parser.add_argument("--url", default=URL, help="Vietlott URL.")
    parser.add_argument(
        "--ajax-key",
        default=os.environ.get("VIETLOTT_535_AJAX_KEY", DEFAULT_AJAX_KEY),
        help="Key trong JS AjaxPro (hoac dat bien VIETLOTT_535_AJAX_KEY).",
    )
    return parser.parse_args()


def _first_result_table_row(soup: BeautifulSoup):
    row = soup.select_one("#divResultContent table tbody tr")
    if row is not None:
        return row
    return soup.select_one("table.table-hover tbody tr") or soup.select_one("table tbody tr")


def _parse_row_element_to_record(row) -> dict:
    cols = row.select("td")
    if len(cols) < 3:
        raise RuntimeError("Khong doc duoc du lieu cot ngay/id/result.")

    date_text = cols[0].get_text(strip=True)
    id_text = cols[1].get_text(strip=True)
    id_num = int(id_text)

    spans = cols[2].select("span")
    values: list[str] = []
    for span in spans:
        txt = span.get_text(strip=True)
        if not txt or txt == "|":
            continue
        if txt.isdigit():
            values.append(str(int(txt)))

    if len(values) < 6:
        raise RuntimeError(f"Khong du 6 so de tao ket qua: {values}")

    nums = values[:5]
    special = values[5]
    result = ",".join(nums) + "|" + special

    return {"date": date_text, "id": id_num, "result": result}


def _github_actions() -> bool:
    return os.environ.get("GITHUB_ACTIONS", "").lower() == "true"


def _mirror_jsonl_urls() -> list[str]:
    out: list[str] = []
    env_u = os.environ.get("VIETLOTT_535_JSONL_URL", "").strip()
    if env_u:
        out.append(env_u)
    for u in DEFAULT_JSONL_URLS:
        if u not in out:
            out.append(u)
    return out


def _record_from_jsonl_row(row: dict) -> dict:
    """Dong JSONL power535: date YYYY-MM-DD, id string, result [5 main + 1 special]."""
    raw_date = str(row.get("date", "")).strip()
    dt = datetime.strptime(raw_date, "%Y-%m-%d")
    date_text = dt.strftime("%d/%m/%Y")
    id_num = int(str(row.get("id", "")).strip())
    nums = row.get("result")
    if not isinstance(nums, list) or len(nums) != 6:
        raise RuntimeError(f"JSONL result khong hop le: {nums!r}")
    main = ",".join(str(int(x)) for x in nums[:5])
    special = str(int(nums[5]))
    return {"date": date_text, "id": id_num, "result": f"{main}|{special}"}


def _fetch_latest_from_mirror_jsonl() -> dict:
    """
    Doc power535.jsonl (mirror cong dong). Co the cham hon trang Vietlott vai gio.
    """
    ua = _headers(None)["User-Agent"]
    last_err: Exception | None = None
    for src in _mirror_jsonl_urls():
        try:
            resp = requests.get(src, timeout=60, headers={"User-Agent": ua})
            resp.raise_for_status()
            rows: list[dict] = []
            for line in resp.text.splitlines():
                line = line.strip()
                if not line:
                    continue
                rows.append(json.loads(line))
            if not rows:
                raise RuntimeError("File JSONL khong co dong nao.")
            latest = max(rows, key=lambda x: int(str(x.get("id", "0")).strip() or "0"))
            return _record_from_jsonl_row(latest)
        except Exception as e:
            last_err = e
    raise last_err if last_err else RuntimeError("Khong co URL JSONL mirror.")


def fetch_latest_record(url: str, ajax_key: str = DEFAULT_AJAX_KEY):
    errors: list[str] = []
    skip_official = os.environ.get("VIETLOTT_535_SKIP_OFFICIAL", "").lower() in (
        "1",
        "true",
        "yes",
    )

    html_fetchers = (
        ("AjaxPro", lambda: _fetch_html_via_ajaxpro(url, ajax_key)),
        ("GET trang", lambda: _fetch_page_html(url)),
    )
    jsonl_fetcher = (
        "JSONL mirror (vietlott-data)",
        lambda: _fetch_latest_from_mirror_jsonl(),
    )

    if skip_official:
        fetch_order = (jsonl_fetcher,)
    elif _github_actions():
        fetch_order = (jsonl_fetcher,) + html_fetchers
    else:
        fetch_order = html_fetchers + (jsonl_fetcher,)

    record: dict | None = None
    for name, getter in fetch_order:
        try:
            if name == jsonl_fetcher[0]:
                record = getter()
                break
            html = getter()
            soup = BeautifulSoup(html, "html.parser")
            row = _first_result_table_row(soup)
            if row is None:
                raise RuntimeError("Khong tim thay dong ket qua trong HTML.")
            record = _parse_row_element_to_record(row)
            break
        except Exception as e:
            errors.append(f"{name}: {e}")

    if record is None:
        raise RuntimeError("Khong lay duoc du lieu. " + " | ".join(errors))
    return record


def load_existing_rows(file_path: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Khong tim thay file: {file_path}")

    wb = load_workbook(file_path, keep_vba=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        date_val = ws.cell(r, 1).value
        id_val = ws.cell(r, 2).value
        result_val = ws.cell(r, 3).value
        if id_val is None:
            continue
        id_digits = "".join(ch for ch in str(id_val) if ch.isdigit())
        if not id_digits:
            continue
        rows.append(
            {
                "date": str(date_val).strip() if date_val is not None else "",
                "id": int(id_digits),
                "result": str(result_val).strip() if result_val is not None else "",
            }
        )
    return wb, ws, rows


def save_rows_to_workbook(wb, ws, rows, file_path: str):
    df = pd.DataFrame(rows, columns=["date", "id", "result"])
    df = df.drop_duplicates(subset="id").sort_values(by="id")
    df["id"] = df["id"].astype(int).astype(str).str.zfill(5)

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = None
        ws.cell(r, 2).value = None
        ws.cell(r, 3).value = None

    for i, row in df.iterrows():
        ws.cell(row=i + 2, column=1, value=row["date"])
        ws.cell(row=i + 2, column=2, value=row["id"])
        ws.cell(row=i + 2, column=3, value=row["result"])

    ws.freeze_panes = "A2"
    center = Alignment(horizontal="center")
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).alignment = center
        ws.cell(r, 2).alignment = center

    for r in range(2, ws.max_row + 1):
        cell = ws[f"C{r}"]
        if cell.value and "|" in str(cell.value):
            left, right = str(cell.value).split("|", 1)
            rich = CellRichText()
            rich.append(TextBlock(InlineFont(rFont="Calibri", sz=11), left + "|"))
            rich.append(
                TextBlock(
                    InlineFont(rFont="Calibri", sz=16, b=True, color="FF006400"),
                    right,
                )
            )
            cell.value = rich

    rebuild_stats_sheets(wb, ws)
    wb.save(file_path)
    print("Saved rows:", len(df))


def rebuild_stats_sheets(wb, ws):
    main_numbers: list[list[int]] = []
    special_numbers: list[int] = []

    for r in range(2, ws.max_row + 1):
        val = ws[f"C{r}"].value
        if not val:
            continue
        text = str(val)
        if "|" not in text:
            continue
        left, right = text.split("|", 1)
        nums = [int(x) for x in left.split(",") if str(x).strip().isdigit()]
        if len(nums) != 5:
            continue
        if not str(right).strip().isdigit():
            continue
        main_numbers.append(nums)
        special_numbers.append(int(right))

    combo_counts = {k: Counter() for k in range(1, 6)}
    for nums in main_numbers:
        nums = sorted(nums)
        for k in range(1, 6):
            for c in itertools.combinations(nums, k):
                combo_counts[k][c] += 1

    special_counter = Counter(special_numbers)

    for name in ["combo_1", "combo_2", "combo_3", "combo_4", "combo_5", "special_freq"]:
        if name in wb.sheetnames:
            del wb[name]

    for k in range(1, 6):
        ws2 = wb.create_sheet(f"combo_{k}")
        ws2.append(["combo", "appear"])
        rows = [
            (",".join(map(str, c)), cnt)
            for c, cnt in combo_counts[k].items()
            if cnt >= 2
        ]
        rows.sort(key=lambda x: x[1], reverse=True)
        for row in rows:
            ws2.append(row)

    ws2 = wb.create_sheet("special_freq")
    ws2.append(["special", "count"])
    rows = sorted(special_counter.items(), key=lambda x: x[1], reverse=True)
    for row in rows:
        ws2.append(row)


def main():
    args = parse_args()
    wb, ws, existing_rows = load_existing_rows(args.file)
    latest = fetch_latest_record(args.url, ajax_key=args.ajax_key)

    existing_ids = {row["id"] for row in existing_rows}
    if latest["id"] in existing_ids:
        print(f"No new draw. Latest ID {latest['id']} da ton tai.")
        return 0

    print(f"Append new draw: {latest['date']} {latest['id']} {latest['result']}")
    existing_rows.append(latest)
    save_rows_to_workbook(wb, ws, existing_rows, args.file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

